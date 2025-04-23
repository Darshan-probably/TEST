# xlsx_to_pdf_converter.py
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
import openpyxl
from pathlib import Path
import re

class XlsxToPdfConverter:
    def __init__(self, config=None):
        """
        Initialize the converter with optional configuration
        
        Args:
            config (dict): Configuration options for PDF generation
        """
        self.config = config or {}
        # Default settings
        self.config.setdefault('page_size', A4)
        self.config.setdefault('margin', 0.5 * inch)
        self.config.setdefault('include_header', True)
        self.config.setdefault('include_footer', True)
        self.config.setdefault('table_style', None)
        self.config.setdefault('preserve_formatting', True)
        
    def _extract_merged_cell_ranges(self, ws):
        """Extract merged cell ranges from a worksheet"""
        merged_ranges = []
        for merged_cell_range in ws.merged_cells.ranges:
            min_row, min_col, max_row, max_col = (
                merged_cell_range.min_row,
                merged_cell_range.min_col,
                merged_cell_range.max_row,
                merged_cell_range.max_col
            )
            merged_ranges.append((min_row, min_col, max_row, max_col))
        return merged_ranges
    
    def _get_cell_value(self, cell):
        """Get cell value handling various types"""
        value = cell.value
        if value is None:
            return ""
        
        # Format dates
        if cell.is_date and value is not None:
            try:
                return value.strftime("%Y-%m-%d")
            except:
                pass
                
        # Format numbers based on cell formatting
        if isinstance(value, (int, float)):
            if "0.00" in cell.number_format:
                return f"{value:.2f}"
            if "0.0" in cell.number_format:
                return f"{value:.1f}"
            if "#,##" in cell.number_format:
                return f"{value:,}"
                
        return str(value)
    
    def _get_table_style(self, merged_ranges):
        """Create table style with merged cells"""
        style = [
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]
        
        # Add styles for merged cells
        for min_row, min_col, max_row, max_col in merged_ranges:
            # Convert from 1-based to 0-based for reportlab
            style.append((
                'SPAN', 
                (min_col-1, min_row-1), 
                (max_col-1, max_row-1)
            ))
            
        if self.config.get('table_style'):
            style.extend(self.config['table_style'])
            
        return TableStyle(style)
    
    def _apply_cell_styles(self, data, ws, table_style):
        """Apply cell specific styling based on Excel formatting"""
        if not self.config.get('preserve_formatting'):
            return table_style
            
        for row_idx, row in enumerate(data):
            for col_idx, _ in enumerate(row):
                # Convert 0-based to 1-based for excel
                cell = ws.cell(row=row_idx+1, column=col_idx+1)
                
                # Check for bold font
                if cell.font and cell.font.bold:
                    table_style.add('FONTNAME', (col_idx, row_idx), (col_idx, row_idx), 'Helvetica-Bold')
                
                # Check for alignment
                if cell.alignment:
                    if cell.alignment.horizontal == 'right':
                        table_style.add('ALIGN', (col_idx, row_idx), (col_idx, row_idx), 'RIGHT')
                    elif cell.alignment.horizontal == 'center':
                        table_style.add('ALIGN', (col_idx, row_idx), (col_idx, row_idx), 'CENTER')
                        
                # Check for background color
                if cell.fill and cell.fill.start_color and cell.fill.start_color.index not in (None, '00000000'):
                    # Convert Excel color to ReportLab (simplified)
                    rgb = cell.fill.start_color.rgb
                    if rgb:
                        try:
                            # Convert ARGB to RGB
                            if len(rgb) == 8:  # ARGB format
                                rgb = rgb[2:]
                            r = int(rgb[0:2], 16) / 255
                            g = int(rgb[2:4], 16) / 255
                            b = int(rgb[4:6], 16) / 255
                            table_style.add('BACKGROUND', (col_idx, row_idx), (col_idx, row_idx), colors.Color(r, g, b))
                        except:
                            pass
                            
        return table_style
    
    def get_header_elements(self, wb):
        """Extract header information from Excel file"""
        elements = []
        
        if not self.config.get('include_header'):
            return elements
            
        styles = getSampleStyleSheet()
        header_style = styles['Heading1']
        
        # Add title from workbook properties if available
        ws = wb.active
        title = wb.properties.title or ws.title
        
        # Check if there are cells that could be part of a header
        potential_header_info = {}
        
        # Check first few rows for header-like content
        for row in range(1, min(8, ws.max_row + 1)):
            for col in range(1, min(5, ws.max_column + 1)):
                cell = ws.cell(row=row, column=col)
                if cell.value:
                    # Look for date patterns
                    value = str(cell.value)
                    if re.search(r'\d{1,4}[-/]\d{1,2}[-/]\d{1,4}', value):
                        potential_header_info['date'] = value
                    
                    # Look for "lot" or "number" keywords
                    if 'lot' in value.lower() or 'number' in value.lower():
                        potential_header_info['lot_number'] = value
                    
                    # Look for name-like patterns
                    if re.search(r'^[A-Z][a-z]+ [A-Z][a-z]+$', value):
                        potential_header_info['name'] = value
        
        # Add title element
        elements.append(Paragraph(title, header_style))
        
        # Add potential header info
        normal_style = styles['Normal']
        for key, value in potential_header_info.items():
            elements.append(Paragraph(f"{key.replace('_', ' ').title()}: {value}", normal_style))
        
        elements.append(Spacer(1, 0.25 * inch))
        return elements
    
    def get_footer_elements(self):
        """Create footer elements"""
        elements = []
        
        if not self.config.get('include_footer'):
            return elements
            
        elements.append(Spacer(1, 0.25 * inch))
        
        styles = getSampleStyleSheet()
        footer_style = styles['Normal']
        footer_style.alignment = 1  # Center alignment
        
        footer_text = self.config.get('footer_text', "Generated by XLSX to PDF Converter")
        elements.append(Paragraph(footer_text, footer_style))
        
        return elements
        
    def convert(self, input_path, output_path):
        """
        Convert Excel file to PDF
        
        Args:
            input_path (str): Path to input Excel file
            output_path (str): Path to output PDF file
            
        Returns:
            str: Path to generated PDF file
        """
        # Ensure directories exist
        output_dir = Path(output_path).parent
        output_dir.mkdir(exist_ok=True, parents=True)
        
        # Open Excel workbook
        wb = openpyxl.load_workbook(input_path, data_only=True)
        ws = wb.active
        
        # Extract merged cell ranges
        merged_ranges = self._extract_merged_cell_ranges(ws)
        
        # Create PDF document
        doc = SimpleDocTemplate(
            output_path,
            pagesize=self.config['page_size'],
            leftMargin=self.config['margin'],
            rightMargin=self.config['margin'],
            topMargin=self.config['margin'],
            bottomMargin=self.config['margin']
        )
        
        # Extract data from worksheet
        data = []
        for row in range(1, ws.max_row + 1):
            row_data = []
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                row_data.append(self._get_cell_value(cell))
            data.append(row_data)
        
        # Create table
        table = Table(data)
        
        # Get basic table style
        table_style = self._get_table_style(merged_ranges)
        
        # Apply cell-specific styles
        table_style = self._apply_cell_styles(data, ws, table_style)
        
        # Apply table style
        table.setStyle(table_style)
        
        # Build PDF
        elements = []
        
        # Add header
        elements.extend(self.get_header_elements(wb))
        
        # Add table
        elements.append(table)
        
        # Add footer
        elements.extend(self.get_footer_elements())
        
        # Build document
        doc.build(elements)
        
        return output_path