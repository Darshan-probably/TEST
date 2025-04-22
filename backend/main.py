#main.py
from fastapi import FastAPI, UploadFile, File, Form, Query, HTTPException, BackgroundTasks
from fastapi.responses import FileResponse, HTMLResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
from typing import Optional, List, Union
import shutil
import os
import uuid
from pathlib import Path
from invoice_processor import InvoiceProcessor
import openpyxl
from config import get_app_setting, API_ENDPOINTS

# Create app
app = FastAPI(title="Invoice Processor")

UPLOAD_DIR = Path(get_app_setting('upload_dir'))
OUTPUT_DIR = Path(get_app_setting('output_dir'))
REACT_BUILD_DIR = Path("../frontend/build")  # Path to React build directory

# Ensure directories exist
for directory in [UPLOAD_DIR, OUTPUT_DIR]:
    directory.mkdir(exist_ok=True)

# Mount React static files
app.mount("/static", StaticFiles(directory=REACT_BUILD_DIR / "static"), name="static")

# Clean up old files periodically
def cleanup_old_files():
    """Remove files older than 1 hour"""
    import time
    current_time = time.time()
    one_hour = 60 * 60
    
    for directory in [UPLOAD_DIR, OUTPUT_DIR]:
        for file_path in directory.glob("*"):
            if file_path.is_file():
                file_age = current_time - file_path.stat().st_mtime
                if file_age > one_hour:
                    file_path.unlink()

def convert_xlsx_to_preview_data(xlsx_path, max_rows=20):
    """
    Convert the first few rows of an Excel file to a JSON-friendly structure
    for preview rendering
    """
    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        ws = wb.active
        
        # Get column headers (using row 1)
        headers = []
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=1, column=col).value
            headers.append(str(cell_value) if cell_value is not None else "")
        
        # Get data rows (limited to max_rows)
        rows = []
        for row in range(2, min(ws.max_row + 1, max_rows + 2)):
            row_data = []
            for col in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=row, column=col).value
                row_data.append(str(cell_value) if cell_value is not None else "")
            rows.append(row_data)
            
        # Include additional metadata
        metadata = {
            "totalRows": ws.max_row,
            "totalColumns": ws.max_column,
            "sheetName": ws.title,
            "previewRows": len(rows)
        }
            
        return {
            "headers": headers,
            "rows": rows,
            "metadata": metadata
        }
    except Exception as e:
        print(f"Error creating preview: {str(e)}")
        return {"error": str(e)}

@app.post("/preview/")
async def preview_invoice_upload(
    file: UploadFile = File(...),
    bags: Optional[str] = Form(None),
    weight: Optional[str] = Form(None),
    min_percent: Optional[str] = Form("-2"),
    max_percent: Optional[str] = Form("2"),
    preserve_wrapping: Optional[str] = Form("true"),
    # Template and document fields
    template_type: Optional[str] = Form("default"),
    date: Optional[str] = Form(None),
    address: Optional[str] = Form(None),
    lot_number: Optional[str] = Form(None),
    name: Optional[str] = Form(None)
):
    """API endpoint to preview processed invoice file"""
    # Convert form values to appropriate types
    bags_int = int(bags) if bags and bags.strip() else None
    weight_float = float(weight) if weight and weight.strip() else None
    min_percent_float = float(min_percent) if min_percent else -2
    max_percent_float = float(max_percent) if max_percent else 2
    preserve_wrapping_bool = preserve_wrapping.lower() == "true" if preserve_wrapping else True
    
    unique_id = str(uuid.uuid4())
    
    try:
        # Validate file extension
        if not file.filename.lower().endswith('.xlsx'):
            raise HTTPException(status_code=400, detail="Only .xlsx files are supported")
        
        # Save uploaded file temporarily
        input_filename = UPLOAD_DIR / f"{unique_id}_preview_{file.filename}"
        output_filename = OUTPUT_DIR / f"{unique_id}_preview_{file.filename}"
        
        # Save uploaded file
        with open(input_filename, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)
        
        # Configure processor
        config = {
            'min_percent': min_percent_float,
            'max_percent': max_percent_float,
            'font_size': 9,  # Fixed font size to 9
            'preserve_wrapping': preserve_wrapping_bool,
            # Store template and document fields in config
            'template_type': template_type,
            'date': date,
            'address': address,
            'lot_number': lot_number,
            'name': name
        }
        
        if bags_int is not None:
            config['bag_count'] = bags_int
        
        if weight_float is not None:
            config['target_weight'] = weight_float
        
        # Process file
        processor = InvoiceProcessor(config)
        output_path = processor.process_file(str(input_filename), str(output_filename))
        
        # Convert to preview data
        preview_data = convert_xlsx_to_preview_data(output_path)
        
        # Include processing configuration in the response
        preview_data["config"] = {
            "bags": bags_int,
            "weight": weight_float,
            "minPercent": min_percent_float,
            "maxPercent": max_percent_float,
            "preserveWrapping": preserve_wrapping_bool,
            "templateType": template_type,
            "date": date,
            "address": address,
            "lotNumber": lot_number,
            "name": name
        }
        
        # Include original and processed file IDs for later downloading
        preview_data["fileInfo"] = {
            "originalFile": input_filename.name,
            "processedFile": output_filename.name
        }
        
        return preview_data
        
    except Exception as e:
        # Handle any errors
        if input_filename.exists():
            try:
                input_filename.unlink()
            except:
                pass
                
        if output_filename.exists():
            try:
                output_filename.unlink()
            except:
                pass
                
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/process/")
async def process_invoice_upload(
    background_tasks: BackgroundTasks,
    file: UploadFile = File(...),
    bags: Optional[str] = Form(None),
    weight: Optional[str] = Form(None),
    min_percent: Optional[str] = Form("-2"),
    max_percent: Optional[str] = Form("2"),
    preserve_wrapping: Optional[str] = Form("true"),
    # Template and document fields
    template_type: Optional[str] = Form("default"),
    date: Optional[str] = Form(None),
    address: Optional[str] = Form(None),
    lot_number: Optional[str] = Form(None),
    name: Optional[str] = Form(None)
):
    """API endpoint to process invoice file and provide download"""
    # Convert form values to appropriate types
    bags_int = int(bags) if bags and bags.strip() else None
    weight_float = float(weight) if weight and weight.strip() else None
    min_percent_float = float(min_percent) if min_percent else -2
    max_percent_float = float(max_percent) if max_percent else 2
    preserve_wrapping_bool = preserve_wrapping.lower() == "true" if preserve_wrapping else True
    
    unique_id = str(uuid.uuid4())
    
    try:
        # Validate file extension
        if not file.filename.lower().endswith('.xlsx'):
            raise HTTPException(status_code=400, detail="Only .xlsx files are supported")
        
        # Save uploaded file
        input_filename = UPLOAD_DIR / f"{unique_id}_{file.filename}"
        output_filename = OUTPUT_DIR / f"{unique_id}_{file.filename.replace('.xlsx', '_processed.xlsx')}"
        
        with open(input_filename, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)
        
        # Configure processor
        config = {
            'min_percent': min_percent_float,
            'max_percent': max_percent_float,
            'font_size': 9,  # Fixed font size to 9
            'preserve_wrapping': preserve_wrapping_bool,
            # Store template and document fields in config
            'template_type': template_type,
            'date': date,
            'address': address,
            'lot_number': lot_number,
            'name': name
        }
        
        if bags_int is not None:
            config['bag_count'] = bags_int
        
        if weight_float is not None:
            config['target_weight'] = weight_float
        
        # Process file
        processor = InvoiceProcessor(config)
        output_path = processor.process_file(str(input_filename), str(output_filename))
        
        # Schedule cleanup task
        background_tasks.add_task(cleanup_old_files)
        
        # Return download URL and file path
        file_name = os.path.basename(output_path)
        download_url = f"/download/{file_name}"
        
        return {
            "download_url": download_url,
            "file_path": str(output_path)  # Include the file path in the response
        }
        
    except Exception as e:
        # Handle any errors
        if input_filename.exists():
            try:
                input_filename.unlink()
            except:
                pass
                
        if output_filename.exists():
            try:
                output_filename.unlink()
            except:
                pass
                
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/download/{file_name}")
async def download_file(file_name: str):
    """Download processed file"""
    file_path = OUTPUT_DIR / file_name
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="File not found")
    
    return FileResponse(
        path=file_path,
        filename=file_name.split('_', 1)[1],  # Remove UUID prefix
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@app.get("/", response_class=HTMLResponse)
async def read_root():
    """Serve the React frontend"""
    with open(REACT_BUILD_DIR / "index.html", "r") as f:
        return f.read()

if __name__ == "__main__":
    import uvicorn
    uvicorn.run("main:app", host="0.0.0.0", port=8080, reload=True)
