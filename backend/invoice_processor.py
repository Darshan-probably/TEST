# invoice_processor.py

import openpyxl
from openpyxl.styles import Alignment, Font
import random
import re
import copy
from datetime import datetime

class InvoiceProcessor:
    """
    Core logic for processing invoice Excel files with weight distributions
    """
    
    def __init__(self, config=None):
        """Initialize with default configuration or custom settings"""
        self.config = {
            'weight_column': 'G',
            'start_row': 8,
            'total_row': 38,
            'bag_count_cell': 'C38',
            'min_percent': -2,
            'max_percent': 2,
            'target_weight': None,
            'font_size': 9,  # Default to 9
            'bag_count': None,
            'preserve_wrapping': True,  # Default to preserving wrapping
            'template_type': 'default',
            'date': None,
            'address': None,
            'lot_number': None,
            'name': None
        }
        
        if config:
            self.config.update(config)
    
    def get_mergecell_value(self, ws, cell_address):
        """Get value from merged cell ranges"""
        cell = ws[cell_address]
        for merged_range in ws.merged_cells.ranges:
            if cell.coordinate in merged_range:
                return ws.cell(merged_range.min_row, merged_range.min_col).value
        return cell.value
    
    def set_mergecell_value(self, ws, cell_address, value):
        """Safely set value in potentially merged cells"""
        cell = ws[cell_address]
        for merged_range in ws.merged_cells.ranges:
            if cell.coordinate in merged_range:
                top_left = ws.cell(merged_range.min_row, merged_range.min_col)
                top_left.value = value
                return
        cell.value = value
    
    def extract_bag_count(self, bag_text):
        """Extract number of bags from text"""
        if not bag_text: 
            return 0
        match = re.search(r'(\d+)\s*[Bb]ags?', str(bag_text))
        return int(match.group(1)) if match else 0
    
    def distribute_weight(self, target_weight, num_items, min_pct=-2, max_pct=2):
        """Distribute weight with random variations"""
        if num_items <= 0: 
            return []
            
        base = target_weight / num_items
        weights = []
        remaining = target_weight
        
        for _ in range(num_items - 1):
            variation = random.uniform(min_pct, max_pct) / 100
            weight = round(base * (1 + variation) * 100) / 100
            weights.append(weight)
            remaining -= weight
        
        # Add last weight to ensure exact total
        weights.append(round(remaining * 100) / 100)
        random.shuffle(weights)
        return weights
    
    def copy_cell_format_and_properties(self, source_cell, target_cell):
        """
        Copy all formatting and properties from source cell to target cell
        including text wrapping and cell dimensions
        """
        # Copy alignment (including text wrapping)
        if source_cell.alignment:
            target_cell.alignment = copy.copy(source_cell.alignment)
        
        # Copy font
        if source_cell.font:
            target_cell.font = copy.copy(source_cell.font)
        
        # Copy border
        if source_cell.border:
            target_cell.border = copy.copy(source_cell.border)
        
        # Copy fill
        if source_cell.fill:
            target_cell.fill = copy.copy(source_cell.fill)
        
        # Copy number format
        target_cell.number_format = source_cell.number_format
        
        # Copy cell dimensions if workbook has them
        ws = source_cell.parent
        col_letter = source_cell.column_letter
        
        # Copy column width if defined
        if col_letter in ws.column_dimensions:
            source_col_dim = ws.column_dimensions[col_letter]
            if target_cell.column_letter in ws.column_dimensions:
                target_col_dim = ws.column_dimensions[target_cell.column_letter]
                # Copy width
                if source_col_dim.width is not None:
                    target_col_dim.width = source_col_dim.width
                # Copy hidden status
                target_col_dim.hidden = source_col_dim.hidden
        
        # Copy row height if defined
        if source_cell.row in ws.row_dimensions:
            source_row_dim = ws.row_dimensions[source_cell.row]
            if target_cell.row in ws.row_dimensions:
                target_row_dim = ws.row_dimensions[target_cell.row]
                # Copy height
                if source_row_dim.height is not None:
                    target_row_dim.height = source_row_dim.height
                # Copy hidden status
                target_row_dim.hidden = source_row_dim.hidden
    
    def copy_row_formatting(self, ws, source_row, target_row):
        """Copy cell formatting between rows with text wrapping preservation"""
        for col in range(1, ws.max_column + 1):
            src = ws.cell(row=source_row, column=col)
            tgt = ws.cell(row=target_row, column=col)
            self.copy_cell_format_and_properties(src, tgt)
    
    def capture_row_dimensions(self, ws, row_index):
        """Capture row dimensions and cell properties for later restoration"""
        dimensions = {}
        
        # Capture row height
        if row_index in ws.row_dimensions:
            dimensions['row_height'] = ws.row_dimensions[row_index].height
        
        # Capture cell properties for each cell in the row
        cell_properties = {}
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_index, column=col)
            col_letter = cell.column_letter
            
            cell_properties[col_letter] = {
                'alignment': copy.copy(cell.alignment) if cell.alignment else None,
                'font': copy.copy(cell.font) if cell.font else None,
                'border': copy.copy(cell.border) if cell.border else None,
                'fill': copy.copy(cell.fill) if cell.fill else None,
                'number_format': cell.number_format,
                'value': cell.value
            }
            
            # Capture column width
            if col_letter in ws.column_dimensions:
                cell_properties[col_letter]['width'] = ws.column_dimensions[col_letter].width
        
        dimensions['cell_properties'] = cell_properties
        return dimensions
    
    def restore_row_dimensions(self, ws, row_index, dimensions):
        """Restore row dimensions and cell properties"""
        # Restore row height
        if 'row_height' in dimensions and dimensions['row_height'] is not None:
            ws.row_dimensions[row_index].height = dimensions['row_height']
        
        # Restore cell properties
        for col_letter, props in dimensions['cell_properties'].items():
            col_idx = openpyxl.utils.column_index_from_string(col_letter)
            cell = ws.cell(row=row_index, column=col_idx)
            
            # Restore cell properties
            if props['alignment']:
                cell.alignment = props['alignment']
            if props['font']:
                cell.font = props['font']
            if props['border']:
                cell.border = props['border']
            if props['fill']:
                cell.fill = props['fill']
            
            cell.number_format = props['number_format']
            
            # We don't restore the value, as it will be set elsewhere
            
            # Restore column width
            if 'width' in props and props['width'] is not None:
                ws.column_dimensions[col_letter].width = props['width']
    
    def insert_rows_preserving_merges(self, ws, insert_at, count):
        """Insert rows while maintaining merged cell integrity and preserving text wrapping"""
        # Capture dimensions of template row to copy formatting from
        template_row_dimensions = self.capture_row_dimensions(ws, insert_at - 1)
        
        # Store all merged cell ranges for later reconstruction
        merged_ranges = []
        for merged_range in list(ws.merged_cells.ranges):
            merged_ranges.append((
                merged_range.min_row, 
                merged_range.max_row, 
                merged_range.min_col, 
                merged_range.max_col
            ))
            if merged_range.min_row >= insert_at:
                # Remove merged ranges that will be shifted
                ws.merged_cells.ranges.remove(merged_range)
        
        # Insert blank rows
        ws.insert_rows(insert_at, count)
        
        # Reconstruct merged cell ranges
        for min_row, max_row, min_col, max_col in merged_ranges:
            if max_row < insert_at:
                # Range is above insertion point - keep as is
                ws.merge_cells(
                    start_row=min_row,
                    end_row=max_row,
                    start_column=min_col,
                    end_column=max_col
                )
            elif min_row >= insert_at:
                # Range is at or below insertion point - shift it
                ws.merge_cells(
                    start_row=min_row + count,
                    end_row=max_row + count,
                    start_column=min_col,
                    end_column=max_col
                )
            else:
                # Range spans the insertion point
                ws.merge_cells(
                    start_row=min_row,
                    end_row=max_row + count,
                    start_column=min_col,
                    end_column=max_col
                )
        
        # Apply preserved formatting to new rows
        for row in range(insert_at, insert_at + count):
            if self.config['preserve_wrapping']:
                # Apply full formatting including cell dimensions
                self.restore_row_dimensions(ws, row, template_row_dimensions)
            else:
                # Apply basic formatting
                self.copy_row_formatting(ws, insert_at - 1, row)
    
    def find_bag_count_cell(self, ws):
        """Find the actual bag count cell after row insertions"""
        # Start with original cell reference
        original_col = openpyxl.utils.column_index_from_string(
            self.config['bag_count_cell'][0]
        )
        
        # Search for "Bags" text in that column near the total_row
        search_range = 5  # Look 5 rows above and below the total_row
        total_row = self.config['total_row']
        
        for row in range(total_row - search_range, total_row + search_range + 1):
            cell_value = ws.cell(row=row, column=original_col).value
            if cell_value and "bag" in str(cell_value).lower():
                return f"{openpyxl.utils.get_column_letter(original_col)}{row}"
        
        # If not found, return the adjusted original cell
        bag_count_cell = self.config['bag_count_cell']
        original_row = int(bag_count_cell[1:])
        
        # Calculate how many rows were added before this cell
        rows_added = 0
        if original_row >= self.config['start_row'] + self.config['total_row']:
            rows_added = self.config['total_row'] - self.config['start_row'] - original_row
            
        return f"{bag_count_cell[0]}{original_row + rows_added}"
    
    def update_document_info(self, ws):
        """Update document information fields if provided in config"""
        # Placeholder implementation - will need to be customized per template
        template_type = self.config['template_type']
        
        # Example: Update date if provided (cell A4 in default template)
        if self.config['date']:
            date_value = self.config['date']
            try:
                # Try to parse as date if it's a string
                if isinstance(date_value, str):
                    date_obj = datetime.strptime(date_value, '%Y-%m-%d')
                    self.set_mergecell_value(ws, 'A4', date_obj)
                else:
                    self.set_mergecell_value(ws, 'A4', date_value)
            except:
                # If parsing fails, use as is
                self.set_mergecell_value(ws, 'A4', date_value)
        
        # Example: Update lot number if provided (cell B4 in default template)
        if self.config['lot_number']:
            self.set_mergecell_value(ws, 'B4', self.config['lot_number'])
        
        # Example: Update name if provided (cell C3 in default template)
        if self.config['name']:
            self.set_mergecell_value(ws, 'C3', self.config['name'])
        
        # Example: Update address if provided (cell C4 in default template)
        if self.config['address']:
            self.set_mergecell_value(ws, 'C4', self.config['address'])
            
        # Note: These are placeholders - actual cell references should be 
        # determined based on the specific template being used
    
    def process_file(self, file_path, output_path=None):
        """Process invoice Excel file with current configuration"""
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        output_path = output_path or file_path.replace('.xlsx', '_processed.xlsx')
        
        # Get or override bag count
        if self.config['bag_count'] is not None:
            bag_count = self.config['bag_count']
        else:
            # Find the correct bag count cell and extract value
            bag_count_cell = self.config['bag_count_cell']
            bag_count = self.extract_bag_count(
                self.get_mergecell_value(ws, bag_count_cell)
            )
        
        # Calculate needed rows and insert if necessary
        available_rows = self.config['total_row'] - self.config['start_row']
        rows_needed = max(0, bag_count - available_rows)
        
        if rows_needed > 0:
            insert_at = self.config['start_row'] + available_rows
            self.insert_rows_preserving_merges(ws, insert_at, rows_needed)
            # Update total_row after insertion
            self.config['total_row'] += rows_needed
        
        # Now that we've potentially modified the rows, find the current bag count cell
        updated_bag_count_cell = self.find_bag_count_cell(ws)
        
        # Get or calculate target weight
        if not self.config['target_weight']:
            self.config['target_weight'] = sum(
                self.get_mergecell_value(ws, f"{self.config['weight_column']}{row}") or 0
                for row in range(self.config['start_row'], self.config['total_row'])
            )
        
        # Clear existing weights
        for row in range(self.config['start_row'], self.config['total_row']):
            self.set_mergecell_value(ws, f"{self.config['weight_column']}{row}", None)
        
        # Generate and apply new weights
        weights = self.distribute_weight(
            self.config['target_weight'],
            bag_count,
            self.config['min_percent'],
            self.config['max_percent']
        )
        
        # Apply new weights with formatting
        for idx, weight in enumerate(weights):
            cell = f"{self.config['weight_column']}{self.config['start_row'] + idx}"
            self.set_mergecell_value(ws, cell, weight)
            
            # Apply formatting
            target_cell = ws[cell]
            target_cell.alignment = Alignment(horizontal='right')
            
            # Use fixed font size of 9
            current_font = target_cell.font
            font = Font(
                name=current_font.name,
                size=self.config['font_size'],
                bold=current_font.bold,
                italic=current_font.italic
            )
            target_cell.font = font
        
        # Update total and bag count
        self.set_mergecell_value(ws, updated_bag_count_cell, f"{bag_count} Bags")
        total_cell = f"{self.config['weight_column']}{self.config['total_row']}"
        self.set_mergecell_value(ws, total_cell, sum(weights))
        
        # Update document info if template and fields are provided
        self.update_document_info(ws)
        
        # Save the result
        wb.save(output_path)
        return output_path