# app/templates/base.py
"""
Base template class for invoice processing
"""
from typing import Dict, Any, Optional, List
from datetime import datetime
import logging

logger = logging.getLogger(__name__)

class InvoiceTemplate:
    """
    Base class for invoice templates
    
    All templates should inherit from this class and implement the required methods
    """
    def __init__(self, config: Optional[Dict[str, Any]] = None):
        """
        Initialize template with optional configuration
        
        Args:
            config: Configuration dictionary to override defaults
        """
        self.config = self.get_default_config()
        
        if config:
            self.config.update(config)
            
        logger.debug(f"Initialized template with config: {self.config}")
    
    def get_default_config(self) -> Dict[str, Any]:
        """
        Get default configuration for this template
        
        Returns:
            Dictionary of default configuration values
        """
        return {
            'weight_column': 'G',
            'start_row': 8,
            'total_row': 38,
            'bag_count_cell': 'C38',
            'font_size': 9,
            'min_percent': -2,
            'max_percent': 2,
            'preserve_wrapping': True,
            'date_cell': 'A4',
            'lot_number_cell': 'B4',
            'name_cell': 'C3',
            'address_cell': 'C4',
        }
    
    def update_document_info(self, ws) -> None:
        """
        Update document information fields
        
        Args:
            ws: Worksheet object
        """
        from app.utils.excel import set_mergecell_value
        
        # Update date if provided
        if self.config.get('date'):
            date_value = self.config['date']
            try:
                # Try to parse as date if it's a string
                if isinstance(date_value, str):
                    date_obj = datetime.strptime(date_value, '%Y-%m-%d')
                    set_mergecell_value(ws, self.config['date_cell'], date_obj)
                else:
                    set_mergecell_value(ws, self.config['date_cell'], date_value)
            except:
                # If parsing fails, use as is
                set_mergecell_value(ws, self.config['date_cell'], date_value)
        
        # Update lot number if provided
        if self.config.get('lot_number'):
            set_mergecell_value(ws, self.config['lot_number_cell'], self.config['lot_number'])
        
        # Update name if provided
        if self.config.get('name'):
            set_mergecell_value(ws, self.config['name_cell'], self.config['name'])
        
        # Update address if provided
        if self.config.get('address'):
            set_mergecell_value(ws, self.config['address_cell'], self.config['address'])
    
    def find_bag_count_cell(self, ws) -> str:
        """
        Find the actual bag count cell after row insertions
        
        Args:
            ws: Worksheet object
            
        Returns:
            Cell address (e.g., 'C38')
        """
        from openpyxl.utils import column_index_from_string, get_column_letter
        
        # Start with original cell reference
        original_col = column_index_from_string(
            self.config['bag_count_cell'][0]
        )
        
        # Search for "Bags" text in that column near the total_row
        search_range = 5  # Look 5 rows above and below the total_row
        total_row = self.config['total_row']
        
        for row in range(total_row - search_range, total_row + search_range + 1):
            cell_value = ws.cell(row=row, column=original_col).value
            if cell_value and "bag" in str(cell_value).lower():
                return f"{get_column_letter(original_col)}{row}"
        
        # If not found, return the adjusted original cell
        bag_count_cell = self.config['bag_count_cell']
        original_row = int(''.join(filter(str.isdigit, bag_count_cell)))
        
        # Calculate how many rows were added before this cell
        rows_added = 0
        if original_row >= self.config['start_row'] + self.config['total_row']:
            rows_added = self.config['total_row'] - self.config['start_row'] - original_row
            
        return f"{bag_count_cell[0]}{original_row + rows_added}"