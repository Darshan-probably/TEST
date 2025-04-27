# app/utils/excel.py
"""
Excel-specific utility functions
"""
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
import re
from typing import Dict, Any, List, Tuple, Optional, Union

def convert_to_preview_data(xlsx_path: str, max_rows: int = 20) -> Dict[str, Any]:
    """
    Convert the first few rows of an Excel file to a JSON-friendly structure
    for preview rendering
    
    Args:
        xlsx_path: Path to Excel file
        max_rows: Maximum number of rows to include in preview
        
    Returns:
        Dictionary with headers, rows, and metadata
    """
    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        ws = wb.active
        
        # Get column headers (using row 1)
        headers = []
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=1, column=col).value
            headers.append(str(cell_value) if cell_value is not None else "")
        
        # Get data rows (limited to max_rows)
        rows = []
        for row in range(2, min(ws.max_row + 1, max_rows + 2)):
            row_data = []
            for col in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=row, column=col).value
                row_data.append(str(cell_value) if cell_value is not None else "")
            rows.append(row_data)
            
        # Include additional metadata
        metadata = {
            "totalRows": ws.max_row,
            "totalColumns": ws.max_column,
            "sheetName": ws.title,
            "previewRows": len(rows)
        }
            
        return {
            "headers": headers,
            "rows": rows,
            "metadata": metadata
        }
    except Exception as e:
        return {"error": str(e)}

def get_mergecell_value(ws, cell_address: str) -> Any:
    """
    Get value from a potentially merged cell
    
    Args:
        ws: Worksheet object
        cell_address: Cell address (e.g., 'A1')
        
    Returns:
        Cell value
    """
    cell = ws[cell_address]
    for merged_range in ws.merged_cells.ranges:
        if cell.coordinate in merged_range:
            return ws.cell(merged_range.min_row, merged_range.min_col).value
    return cell.value

def set_mergecell_value(ws, cell_address: str, value: Any) -> None:
    """
    Safely set value in potentially merged cells
    
    Args:
        ws: Worksheet object
        cell_address: Cell address (e.g., 'A1')
        value: Value to set
    """
    cell = ws[cell_address]
    for merged_range in ws.merged_cells.ranges:
        if cell.coordinate in merged_range:
            top_left = ws.cell(merged_range.min_row, merged_range.min_col)
            top_left.value = value
            return
    cell.value = value

def extract_bag_count(bag_text: str) -> int:
    """
    Extract number of bags from text
    
    Args:
        bag_text: Text containing bag count (e.g., '10 Bags')
        
    Returns:
        Number of bags as integer
    """
    if not bag_text:
        return 0
    
    match = re.search(r'(\d+)\s*[Bb]ags?', str(bag_text))
    return int(match.group(1)) if match else 0

def capture_row_dimensions(ws, row_index: int) -> Dict[str, Any]:
    """
    Capture row dimensions and cell properties for later restoration
    
    Args:
        ws: Worksheet object
        row_index: Row index
        
    Returns:
        Dictionary of row dimensions and cell properties
    """
    dimensions = {}
    
    # Capture row height
    if row_index in ws.row_dimensions:
        dimensions['row_height'] = ws.row_dimensions[row_index].height
    
    # Capture cell properties for each cell in the row
    cell_properties = {}
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=row_index, column=col)
        col_letter = cell.column_letter
        
        cell_properties[col_letter] = {
            'alignment': cell.alignment and cell.alignment,
            'font': cell.font and cell.font,
            'border': cell.border and cell.border,
            'fill': cell.fill and cell.fill,
            'number_format': cell.number_format,
            'value': cell.value
        }
        
        # Capture column width
        if col_letter in ws.column_dimensions:
            cell_properties[col_letter]['width'] = ws.column_dimensions[col_letter].width
    
    dimensions['cell_properties'] = cell_properties
    return dimensions

def restore_row_dimensions(ws, row_index: int, dimensions: Dict[str, Any]) -> None:
    """
    Restore row dimensions and cell properties
    
    Args:
        ws: Worksheet object
        row_index: Row index
        dimensions: Dictionary of row dimensions and cell properties
    """
    # Restore row height
    if 'row_height' in dimensions and dimensions['row_height'] is not None:
        ws.row_dimensions[row_index].height = dimensions['row_height']
    
    # Restore cell properties
    for col_letter, props in dimensions['cell_properties'].items():
        col_idx = column_index_from_string(col_letter)
        cell = ws.cell(row=row_index, column=col_idx)
        
        # Restore cell properties
        if props['alignment']:
            cell.alignment = props['alignment']
        if props['font']:
            cell.font = props['font']
        if props['border']:
            cell.border = props['border']
        if props['fill']:
            cell.fill = props['fill']
        
        cell.number_format = props['number_format']
        
        # We don't restore the value, as it will be set elsewhere
        
        # Restore column width
        if 'width' in props and props['width'] is not None:
            ws.column_dimensions[col_letter].width = props['width']