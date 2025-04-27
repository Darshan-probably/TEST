# app/processing/invoice_processor.py
"""
Excel invoice processor for weight distribution and formatting
"""
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from pathlib import Path
from typing import Dict, Any, Optional, Union
import logging

from app.processing.weight_distribution import WeightDistributor
from app.processing.formatter import CellFormatter
from app.utils.excel import extract_bag_count, set_mergecell_value, get_mergecell_value
from app.templates.default_template import DEFAULT_TEMPLATE
from app.templates.custom_template import CUSTOM_TEMPLATE

logger = logging.getLogger(__name__)

class InvoiceProcessor:
    """
    Handles processing of invoice Excel files including weight distribution
    """
    def __init__(self, config: Optional[Dict[str, Any]] = None):
        """
        Initialize processor with configuration
        
        Args:
            config: Configuration dictionary
        """
        self.config = DEFAULT_TEMPLATE.copy()
        
        if config:
            self.config.update({k: v for k, v in config.items() if v is not None})
            
        # Initialize weight distributor
        self.weight_distributor = WeightDistributor(
            min_percent=self.config.get('min_percent', -2),
            max_percent=self.config.get('max_percent', 2)
        )
        
        logger.debug(f"Initialized invoice processor with config: {self.config}")

    def process_file(self, input_path: str, output_path: str) -> str:
        """
        Process invoice Excel file
        
        Args:
            input_path: Path to input file
            output_path: Path to output file
            
        Returns:
            Path to processed file
        """
        input_path = Path(input_path)
        output_path = Path(output_path)
        
        logger.info(f"Processing invoice: {input_path} -> {output_path}")
        
        # Load workbook
        wb = openpyxl.load_workbook(input_path)
        ws = wb.active
        
        # Update document information
        self._update_document_info(ws)
        
        # Process weight distribution if bag count and target weight are provided
        bag_count = self.config.get('bag_count')
        target_weight = self.config.get('target_weight')
        
        if bag_count is not None and target_weight is not None:
            logger.info(f"Distributing {target_weight} across {bag_count} bags")
            self._distribute_weights(ws, bag_count, target_weight)
        elif bag_count is not None or target_weight is not None:
            logger.warning("Both bag_count and target_weight must be provided for weight distribution")
        
        # Save workbook
        wb.save(output_path)
        logger.info(f"Saved processed file to {output_path}")
        
        return str(output_path)
        
    def _update_document_info(self, ws) -> None:
        """
        Update document information fields
        
        Args:
            ws: Worksheet object
        """
        # Update date if provided
        if self.config.get('date'):
            set_mergecell_value(ws, self.config['date_cell'], self.config['date'])
        
        # Update lot number if provided
        if self.config.get('lot_number'):
            set_mergecell_value(ws, self.config['lot_number_cell'], self.config['lot_number'])
        
        # Update name if provided
        if self.config.get('name'):
            set_mergecell_value(ws, self.config['name_cell'], self.config['name'])
        
        # Update address if provided
        if self.config.get('address'):
            set_mergecell_value(ws, self.config['address_cell'], self.config['address'])
            
        # Set bag count in the appropriate cell if provided
        if self.config.get('bag_count'):
            self._set_bag_count(ws, self.config['bag_count'])
        
    def _set_bag_count(self, ws, bag_count: int) -> None:
        """
        Set bag count in the worksheet
        
        Args:
            ws: Worksheet object
            bag_count: Number of bags
        """
        # Find the bag count cell
        bag_count_cell = self.config['bag_count_cell']
        
        # Try to extract any existing bag count
        current_value = get_mergecell_value(ws, bag_count_cell)
        
        # If there's an existing value, replace the number but keep the text format
        if current_value:
            if isinstance(current_value, str) and 'bag' in current_value.lower():
                # Replace just the number in the string
                import re
                new_value = re.sub(r'\d+', str(bag_count), current_value)
                set_mergecell_value(ws, bag_count_cell, new_value)
            else:
                # Just set the bag count with "Bags" text
                set_mergecell_value(ws, bag_count_cell, f"{bag_count} Bags")
        else:
            # No existing value, just set the bag count with "Bags" text
            set_mergecell_value(ws, bag_count_cell, f"{bag_count} Bags")
            
    def _distribute_weights(self, ws, bag_count: int, target_weight: float) -> None:
        """
        Distribute weights across invoice rows
        
        Args:
            ws: Worksheet object
            bag_count: Number of bags
            target_weight: Target total weight
        """
        # Get configuration values
        weight_col = column_index_from_string(self.config['weight_column'])
        start_row = self.config['start_row']
        total_row = self.config['total_row']
        font_size = self.config.get('font_size', 9)
        preserve_wrapping = self.config.get('preserve_wrapping', True)
        
        # Determine if we need to insert more rows
        required_rows = bag_count
        available_rows = total_row - start_row
        
        if required_rows > available_rows:
            # Calculate how many rows to insert
            rows_to_insert = required_rows - available_rows
            logger.info(f"Inserting {rows_to_insert} rows to accommodate {bag_count} bags")
            
            if preserve_wrapping:
                # Use formatter to insert rows while preserving formatting
                template_row = start_row  # Use the first row as template
                CellFormatter.insert_rows_preserving_merges(
                    ws, 
                    insert_at=start_row + 1,  # Insert after the first row
                    count=rows_to_insert, 
                    template_row=template_row
                )
            else:
                # Simple row insertion without preserving formatting
                ws.insert_rows(start_row + 1, rows_to_insert)
        
        # Generate distributed weights
        weights = self.weight_distributor.distribute_weight(target_weight, bag_count)
        
        # Apply weights to cells
        for i, weight in enumerate(weights):
            row = start_row + i
            cell = ws.cell(row=row, column=weight_col)
            cell.value = weight
            
            # Format the cell
            cell_address = f"{get_column_letter(weight_col)}{row}"
            CellFormatter.apply_weight_cell_formatting(ws, cell_address, font_size)
        
        # Clear any excess rows if needed
        if available_rows > required_rows:
            for row in range(start_row + required_rows, total_row):
                ws.cell(row=row, column=weight_col).value = None
        
        # Update total formula if needed
        total_cell = ws.cell(row=total_row, column=weight_col)
        
        # Set or update the total cell
        if total_cell.value is None or not str(total_cell.value).startswith('='):
            # No formula exists, just set the sum directly
            total_cell.value = target_weight
        else:
            # Adjust cell formula to match the new row count
            new_formula = f"=SUM({get_column_letter(weight_col)}{start_row}:{get_column_letter(weight_col)}{start_row + bag_count - 1})"
            total_cell.value = new_formula