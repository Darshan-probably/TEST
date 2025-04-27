# app/processing/pdf_converter.py
"""
Excel to PDF conversion functionality
"""
import os
from pathlib import Path
from typing import Dict, Any, Optional, Union
import logging
import subprocess
import platform
from reportlab.lib.pagesizes import letter, A4
from reportlab.pdfgen import canvas
import openpyxl
import tempfile

from app.core.exceptions import ProcessingError

logger = logging.getLogger(__name__)

class XlsxToPdfConverter:
    """
    Handles conversion of Excel files to PDF format
    """
    def __init__(self, config: Optional[Dict[str, Any]] = None):
        """
        Initialize converter with configuration
        
        Args:
            config: Configuration options
        """
        self.config = {
            'include_header': True,
            'include_footer': True,
            'preserve_formatting': True,
            'page_size': A4,
            'footer_text': None
        }
        
        if config:
            self.config.update({k: v for k, v in config.items() if v is not None})
        
        logger.debug(f"Initialized PDF converter with config: {self.config}")
    
    def convert(self, input_path: str, output_path: str) -> Path:
        """
        Convert Excel file to PDF
        
        This method attempts multiple conversion strategies depending on the platform.
        1. First, try using specialized libraries if available (depends on OS)
        2. If that fails, fall back to a basic PDF generation using ReportLab
        
        Args:
            input_path: Path to Excel file
            output_path: Path where PDF should be saved
            
        Returns:
            Path to the generated PDF file
        """
        input_path = Path(input_path)
        output_path = Path(output_path)
        
        if not input_path.exists():
            raise FileNotFoundError(f"Input file not found: {input_path}")
        
        logger.info(f"Converting Excel to PDF: {input_path} -> {output_path}")
        
        # Try platform-specific methods first
        try:
            if self._try_platform_specific_conversion(input_path, output_path):
                return output_path
        except Exception as e:
            logger.warning(f"Platform-specific conversion failed: {e}")
        
        # Fall back to basic conversion
        try:
            self._basic_conversion(input_path, output_path)
            return output_path
        except Exception as e:
            raise ProcessingError(f"PDF conversion failed: {e}")
    
    def _try_platform_specific_conversion(self, input_path: Path, output_path: Path) -> bool:
        """
        Attempt platform-specific conversion which typically preserves formatting better
        
        Args:
            input_path: Input Excel file path
            output_path: Output PDF file path
            
        Returns:
            True if conversion succeeded
        """
        system = platform.system().lower()
        
        if system == 'windows':
            return self._convert_using_excel_com(input_path, output_path)
        elif system == 'darwin':  # macOS
            return self._convert_using_applescript(input_path, output_path)
        elif system == 'linux':
            return self._convert_using_libreoffice(input_path, output_path)
        
        return False
    
    def _convert_using_applescript(self, input_path: Path, output_path: Path) -> bool:
        """
        Convert using AppleScript on macOS
        
        Args:
            input_path: Input Excel file path
            output_path: Output PDF file path
            
        Returns:
            True if conversion succeeded
        """
        script = f'''
        tell application "Microsoft Excel"
            open "{input_path.absolute()}"
            set theDoc to active workbook
            save workbook as theDoc filename "{output_path.absolute()}" file format PDF file format
            close theDoc saving no
            quit
        end tell
        '''
        
        try:
            import subprocess
            subprocess.run(['osascript', '-e', script], check=True)
            return True
        except Exception as e:
            logger.error(f"AppleScript conversion failed: {e}")
            return False
    
    def _convert_using_libreoffice(self, input_path: Path, output_path: Path) -> bool:
        """
        Convert using LibreOffice on Linux
        
        Args:
            input_path: Input Excel file path
            output_path: Output PDF file path
            
        Returns:
            True if conversion succeeded
        """
        try:
            cmd = [
                'libreoffice', '--headless', '--convert-to', 'pdf',
                '--outdir', str(output_path.parent.absolute()),
                str(input_path.absolute())
            ]
            subprocess.run(cmd, check=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
            
            # LibreOffice creates the PDF with the same basename as the input file
            # so we may need to rename it
            expected_output = output_path.parent / f"{input_path.stem}.pdf"
            if expected_output != output_path and expected_output.exists():
                os.rename(expected_output, output_path)
                
            return output_path.exists()
        except Exception as e:
            logger.error(f"LibreOffice conversion failed: {e}")
            return False
    
    def _basic_conversion(self, input_path: Path, output_path: Path) -> None:
        """
        Basic conversion using ReportLab
        This is a fallback method that doesn't preserve all formatting
        
        Args:
            input_path: Input Excel file path
            output_path: Output PDF file path
        """
        logger.info("Using basic conversion method with ReportLab")
        
        # Load the Excel file
        wb = openpyxl.load_workbook(input_path, data_only=True)
        ws = wb.active
        
        # Configuration for PDF
        page_size = self.config.get('page_size', A4)
        margin = 50  # points
        max_rows_per_page = 40
        
        # Create PDF
        c = canvas.Canvas(str(output_path), pagesize=page_size)
        width, height = page_size
        
        # Calculate usable space
        usable_width = width - 2 * margin
        usable_height = height - 2 * margin
        
        # Add header if requested
        if self.config.get('include_header', True):
            self._add_header(c, width, height, input_path.name)
        
        # Determine column widths
        col_widths = self._calculate_column_widths(ws, usable_width)
        
        # Draw content
        row_height = 20  # points
        y_position = height - margin - 40  # Start below header
        
        rows_on_current_page = 0
        for row_idx in range(1, ws.max_row + 1):
            # Check if we need a new page
            if rows_on_current_page >= max_rows_per_page:
                if self.config.get('include_footer', True):
                    self._add_footer(c, width, height)
                c.showPage()
                y_position = height - margin - 40
                rows_on_current_page = 0
                if self.config.get('include_header', True):
                    self._add_header(c, width, height, input_path.name)
                    y_position -= 40
            
            # Draw row
            x_position = margin
            for col_idx in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if cell_value is not None:
                    c.drawString(x_position, y_position, str(cell_value))
                x_position += col_widths[col_idx - 1]
            
            y_position -= row_height
            rows_on_current_page += 1
        
        # Add footer to the last page
        if self.config.get('include_footer', True):
            self._add_footer(c, width, height)
        
        c.save()
    
    def _add_header(self, canvas_obj, width: float, height: float, filename: str) -> None:
        """
        Add header to PDF page
        
        Args:
            canvas_obj: ReportLab Canvas object
            width: Page width
            height: Page height
            filename: Original filename
        """
        canvas_obj.setFont("Helvetica-Bold", 12)
        canvas_obj.drawString(50, height - 40, f"File: {filename}")
        canvas_obj.line(50, height - 50, width - 50, height - 50)
    
    def _add_footer(self, canvas_obj, width: float, height: float) -> None:
        """
        Add footer to PDF page
        
        Args:
            canvas_obj: ReportLab Canvas object
            width: Page width
            height: Page height
        """
        footer_text = self.config.get('footer_text', "Generated by Excel Processor")
        
        canvas_obj.setFont("Helvetica", 9)
        canvas_obj.drawString(50, 40, footer_text)
        canvas_obj.drawString(width - 120, 40, "Page: " + str(canvas_obj.getPageNumber()))
        canvas_obj.line(50, 50, width - 50, 50)
    
    def _calculate_column_widths(self, ws, usable_width: float) -> list:
        """
        Calculate column widths proportionally
        
        Args:
            ws: Worksheet object
            usable_width: Usable width on page
            
        Returns:
            List of column widths
        """
        # Get column dimensions
        col_dimensions = []
        for col_idx in range(1, ws.max_column + 1):
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            col_width = ws.column_dimensions[col_letter].width if col_letter in ws.column_dimensions else 8.43  # Default width
            col_dimensions.append(col_width)
        
        # Calculate proportional widths
        total_excel_width = sum(col_dimensions)
        col_widths = [usable_width * dim / total_excel_width for dim in col_dimensions]
        
        return col_widths