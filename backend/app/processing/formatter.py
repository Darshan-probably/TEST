# app/processing/formatter.py
"""
Cell formatting utilities for Excel processing
"""
import copy
import openpyxl
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from typing import Dict, Any, List, Tuple, Optional
import logging

logger = logging.getLogger(__name__)

class CellFormatter:
    """
    Handles copying and applying formatting to Excel cells
    """
    @staticmethod
    def copy_cell_format_and_properties(source_cell, target_cell) -> None:
        """
        Copy all formatting and properties from source cell to target cell
        including text wrapping and cell dimensions
        
        Args:
            source_cell: Source cell to copy from
            target_cell: Target cell to copy to
        """
        # Copy alignment (including text wrapping)
        if source_cell.alignment:
            target_cell.alignment = copy.copy(source_cell.alignment)
        
        # Copy font
        if source_cell.font:
            target_cell.font = copy.copy(source_cell.font)
        
        # Copy border
        if source_cell.border:
            target_cell.border = copy.copy(source_cell.border)
        
        # Copy fill
        if source_cell.fill:
            target_cell.fill = copy.copy(source_cell.fill)
        
        # Copy number format
        target_cell.number_format = source_cell.number_format
        
        # Copy cell dimensions if workbook has them
        ws = source_cell.parent
        col_letter = source_cell.column_letter
        
        # Copy column width if defined
        if col_letter in ws.column_dimensions:
            source_col_dim = ws.column_dimensions[col_letter]
            if target_cell.column_letter in ws.column_dimensions:
                target_col_dim = ws.column_dimensions[target_cell.column_letter]
                # Copy width
                if source_col_dim.width is not None:
                    target_col_dim.width = source_col_dim.width
                # Copy hidden status
                target_col_dim.hidden = source_col_dim.hidden
        
        # Copy row height if defined
        if source_cell.row in ws.row_dimensions:
            source_row_dim = ws.row_dimensions[source_cell.row]
            if target_cell.row in ws.row_dimensions:
                target_row_dim = ws.row_dimensions[target_cell.row]
                # Copy height
                if source_row_dim.height is not None:
                    target_row_dim.height = source_row_dim.height
                # Copy hidden status
                target_row_dim.hidden = source_row_dim.hidden
    
    @staticmethod
    def copy_row_formatting(ws, source_row: int, target_row: int) -> None:
        """
        Copy cell formatting between rows with text wrapping preservation
        
        Args:
            ws: Worksheet object
            source_row: Source row to copy from
            target_row: Target row to copy to
        """
        for col in range(1, ws.max_column + 1):
            src = ws.cell(row=source_row, column=col)
            tgt = ws.cell(row=target_row, column=col)
            CellFormatter.copy_cell_format_and_properties(src, tgt)
    
    @staticmethod
    def insert_rows_preserving_merges(ws, insert_at: int, count: int, template_row: int) -> None:
        """
        Insert rows while maintaining merged cell integrity and preserving text wrapping
        
        Args:
            ws: Worksheet object
            insert_at: Row index to insert at
            count: Number of rows to insert
            template_row: Row to use as template for formatting
        """
        from app.utils.excel import capture_row_dimensions, restore_row_dimensions
        
        # Capture dimensions of template row to copy formatting from
        template_row_dimensions = capture_row_dimensions(ws, template_row)
        
        # Store all merged cell ranges for later reconstruction
        merged_ranges = []
        for merged_range in list(ws.merged_cells.ranges):
            merged_ranges.append((
                merged_range.min_row, 
                merged_range.max_row, 
                merged_range.min_col, 
                merged_range.max_col
            ))
            if merged_range.min_row >= insert_at:
                # Remove merged ranges that will be shifted
                ws.merged_cells.ranges.remove(merged_range)
        
        # Insert blank rows
        ws.insert_rows(insert_at, count)
        
        # Reconstruct merged cell ranges
        for min_row, max_row, min_col, max_col in merged_ranges:
            if max_row < insert_at:
                # Range is above insertion point - keep as is
                ws.merge_cells(
                    start_row=min_row,
                    end_row=max_row,
                    start_column=min_col,
                    end_column=max_col
                )
            elif min_row >= insert_at:
                # Range is at or below insertion point - shift it
                ws.merge_cells(
                    start_row=min_row + count,
                    end_row=max_row + count,
                    start_column=min_col,
                    end_column=max_col
                )
            else:
                # Range spans the insertion point
                ws.merge_cells(
                    start_row=min_row,
                    end_row=max_row + count,
                    start_column=min_col,
                    end_column=max_col
                )
        
        # Apply preserved formatting to new rows
        for row in range(insert_at, insert_at + count):
            # Apply full formatting including cell dimensions
            restore_row_dimensions(ws, row, template_row_dimensions)
        
        logger.debug(f"Inserted {count} rows at row {insert_at} with formatting from row {template_row}")
    
    @staticmethod
    def apply_weight_cell_formatting(ws, cell_address: str, font_size: int = 9) -> None:
        """
        Apply standard formatting to a weight cell
        
        Args:
            ws: Worksheet object
            cell_address: Cell address (e.g., 'G10')
            font_size: Font size for the cell
        """
        # Get the cell
        cell = ws[cell_address]
        
        # Right-align the cell
        cell.alignment = Alignment(horizontal='right')
        
        # Apply font formatting
        current_font = cell.font or Font()
        cell.font = Font(
            name=current_font.name or 'Calibri',
            size=font_size,
            bold=current_font.bold,
            italic=current_font.italic
        )